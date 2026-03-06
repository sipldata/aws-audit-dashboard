"""
AWS Internal Audit - Excel Report Generator
Matches the exact internal checklist sheet format:
  Sheet 1: AWS Account Structure Sheet (master — all accounts)
  Sheet 2: Bills — Service x Region breakdown
  Sheet 3: Region & Service Tracking (all resources per region)
  Sheet 4: Service Quotas (EC2 On-Demand Standard, G&VT, Elastic IPs)
  Sheet 5: Organization Verification
  Sheet 6: Checklist Status (6-step completion tracker)
"""

import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────
DARK_BLUE = "1B2A4A"
MID_BLUE = "2F5496"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=WHITE)
TITLE_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type="solid")
SUB_FONT = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
SUB_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
NORMAL = Font(name="Calibri", size=10)
BOLD = Font(name="Calibri", bold=True, size=10)
ERR_FONT = Font(name="Calibri", color="E74C3C", bold=True)
OK_FONT = Font(name="Calibri", color="27AE60", bold=True)
WARN_FONT = Font(name="Calibri", color="F39C12", bold=True)
ALT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"),
)


def _widths(ws, w):
    for c, width in w.items():
        ws.column_dimensions[get_column_letter(c)].width = width


def _title(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = CENTER
    ws.row_dimensions[row].height = 30
    return row + 1


def _headers(ws, row, cols):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 22
    return row + 1


def _row(ws, row, vals, alt=False):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = NORMAL; c.alignment = LEFT; c.border = BORDER
        if alt:
            c.fill = ALT_FILL
    return row + 1


def _status_font(ws, row, col, val):
    v = str(val).lower()
    cell = ws.cell(row=row, column=col)
    if v in ("enabled", "ok", "yes", "done", "running", "available"):
        cell.font = OK_FONT
    elif v in ("disabled", "error", "no", "alarm", "stopped") or v.startswith("error"):
        cell.font = ERR_FONT
    elif v in ("pending", "warning", "due"):
        cell.font = WARN_FONT


def generate_report(audit_results, save_dir=None, year=None, month=None):
    wb = Workbook()
    now = datetime.utcnow()
    ry = year or now.year
    rm = month or now.month

    # ==================================================================
    # SHEET 1: AWS Account Structure Sheet (Master)
    # Columns match the internal checklist fields exactly
    # ==================================================================
    ws = wb.active
    ws.title = "AWS Account Structure"
    ws.sheet_properties.tabColor = DARK_BLUE

    cols = [
        "#",                                    # 1
        "Account Label / Client",               # 2
        "Account ID",                           # 3
        "MFA Status",                           # 4
        "Last Month Total Cost (USD)",          # 5  Step 1
        "Forecasted Cost Current Month (USD)",  # 6  Step 1
        "Outstanding Balance (USD)",            # 7  Step 2
        "Bill Due",                             # 8  Step 2
        "Active Regions",                       # 9  Step 3
        "Services Running",                     # 10 Step 3
        "Total Resources",                      # 11 Step 3
        "Organization ID",                      # 12 Step 4
        "Management Account Email",             # 13 Step 4
        "On-Demand Std (N.Virginia)",           # 14 Step 5
        "G&VT Instances (N.Virginia)",          # 15 Step 5
        "Elastic IPs (N.Virginia)",             # 16 Step 5
        "On-Demand Std (Mumbai)",               # 17 Step 5
        "G&VT Instances (Mumbai)",              # 18 Step 5
        "Elastic IPs (Mumbai)",                 # 19 Step 5
        "On-Demand Std (Hyderabad)",            # 20 Step 5
        "On-Demand Std (Singapore)",            # 21 Step 5
        "Last Updated",                         # 22 Step 6
        "Remarks / Errors",                     # 23
    ]
    ncols = len(cols)

    row = _title(ws, 1, "AWS INTERNAL AUDIT - ACCOUNT STRUCTURE SHEET", ncols)
    row = _title(ws, row,
                 f"Period: {datetime(ry, rm, 1).strftime('%B %Y')} | "
                 f"Generated: {now.strftime('%Y-%m-%d %H:%M UTC')}", ncols)
    ws.cell(row=row-1, column=1).font = Font(name="Calibri", size=10, color=WHITE, italic=True)
    row = _headers(ws, row, cols)

    for idx, audit in enumerate(audit_results, 1):
        b = audit.get("billing") or {}
        p = audit.get("payment") or {}
        org = audit.get("organization") or {}
        rd = audit.get("regions") or {}
        q = audit.get("quotas") or {}
        errs = audit.get("errors") or []

        ar = rd.get("regions", {}) if isinstance(rd, dict) else {}
        total_res = rd.get("total_resources", 0) if isinstance(rd, dict) else 0

        # Gather service names across all regions
        all_svcs = set()
        for rdata in ar.values():
            all_svcs.update(rdata.get("services", {}).keys())

        region_list = ", ".join(sorted(ar.keys())) if ar else "None"
        svc_list = ", ".join(sorted(all_svcs)) if all_svcs else "None"

        def _qv(region_key, qname):
            for rk, rv in q.items():
                if region_key in rk:
                    qd = rv.get(qname, {})
                    if "error" in qd:
                        return "Error"
                    return qd.get("value", "N/A")
            return "N/A"

        vals = [
            idx,
            audit.get("account_label", ""),
            audit.get("account_id", ""),
            org.get("mfa_status", "N/A"),
            b.get("last_month_total_cost", "N/A"),
            b.get("forecasted_cost_current_month", "N/A"),
            p.get("outstanding_balance", "N/A"),
            p.get("payment_due", "No"),
            region_list,
            svc_list,
            total_res,
            org.get("org_id", "N/A"),
            org.get("management_account_email", "N/A"),
            _qv("us-east-1", "Running On-Demand Standard (A,C,D,H,I,M,R,T,Z) Instances"),
            _qv("us-east-1", "Running On-Demand G and VT Instances"),
            _qv("us-east-1", "EC2-VPC Elastic IPs"),
            _qv("ap-south-1", "Running On-Demand Standard (A,C,D,H,I,M,R,T,Z) Instances"),
            _qv("ap-south-1", "Running On-Demand G and VT Instances"),
            _qv("ap-south-1", "EC2-VPC Elastic IPs"),
            _qv("ap-south-2", "Running On-Demand Standard (A,C,D,H,I,M,R,T,Z) Instances"),
            _qv("ap-southeast-1", "Running On-Demand Standard (A,C,D,H,I,M,R,T,Z) Instances"),
            audit.get("last_updated", audit.get("audit_timestamp", "")),
            "; ".join(errs) if errs else "OK",
        ]

        row = _row(ws, row, vals, alt=(idx % 2 == 0))

        # Color-code key cells
        _status_font(ws, row-1, 4, org.get("mfa_status", ""))    # MFA
        _status_font(ws, row-1, 8, p.get("payment_due", "No"))    # Bill Due
        if errs:
            ws.cell(row=row-1, column=23).font = ERR_FONT
        else:
            ws.cell(row=row-1, column=23).font = OK_FONT

    _widths(ws, {
        1:4, 2:22, 3:16, 4:12, 5:20, 6:24, 7:20, 8:10,
        9:40, 10:50, 11:12, 12:18, 13:30,
        14:22, 15:22, 16:18, 17:22, 18:22, 19:18,
        20:22, 21:22, 22:22, 23:35,
    })
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{row-1}"
    ws.freeze_panes = "A4"

    # ==================================================================
    # SHEET 2: Bills — Service x Region (from Billing > Bills page)
    # ==================================================================
    ws_bills = wb.create_sheet(title="Bills - Service x Region")
    ws_bills.sheet_properties.tabColor = "E74C3C"
    bcols = ["Account", "Period", "Service", "Region", "Cost (USD)"]
    bncols = len(bcols)

    row = _title(ws_bills, 1, "BILLING > BILLS - SERVICE BY REGION BREAKDOWN", bncols)
    row = _headers(ws_bills, row, bcols)

    for audit in audit_results:
        label = audit.get("account_label", "")
        bills = audit.get("bills") or {}

        for period_label, period_name in [("last_month", "Last Month"), ("current_month", "Current Month")]:
            items = bills.get(period_label, [])
            if not items:
                row = _row(ws_bills, row, [label, period_name, "No data", "", ""])
                continue
            first = True
            for item in sorted(items, key=lambda x: float(x.get("Cost (USD)", "0")), reverse=True):
                row = _row(ws_bills, row, [
                    label if first else "",
                    period_name if first else "",
                    item.get("Service", ""),
                    item.get("Region", ""),
                    item.get("Cost (USD)", ""),
                ], alt=(row % 2 == 0))
                first = False

    _widths(ws_bills, {1:22, 2:16, 3:45, 4:20, 5:14})
    ws_bills.auto_filter.ref = f"A2:{get_column_letter(bncols)}{row-1}"
    ws_bills.freeze_panes = "A3"

    # ==================================================================
    # SHEET 3: Region & Service Tracking (every resource)
    # ==================================================================
    ws_reg = wb.create_sheet(title="Region & Services")
    ws_reg.sheet_properties.tabColor = "27AE60"
    rcols = ["Account", "Region", "Region Name", "Service", "Count", "Resource Details"]
    rncols = len(rcols)

    row = _title(ws_reg, 1, "REGION & SERVICE TRACKING - ALL RESOURCES", rncols)
    row = _headers(ws_reg, row, rcols)

    for audit in audit_results:
        label = audit.get("account_label", "")
        rd = audit.get("regions") or {}
        ar = rd.get("regions", {}) if isinstance(rd, dict) else {}

        if not ar:
            row = _row(ws_reg, row, [label, "No resources", "", "", "", ""])
            continue

        first_acct = True
        for region, data in sorted(ar.items()):
            svcs = data.get("services", {})
            details = data.get("details", {})
            rname = data.get("region_name", region)
            first_r = True
            for svc, cnt in svcs.items():
                items = details.get(svc, [])
                detail_lines = []
                for item in items[:20]:
                    detail_lines.append(" | ".join(f"{k}: {v}" for k, v in item.items()))
                dtxt = "\n".join(detail_lines)
                if len(items) > 20:
                    dtxt += f"\n... +{len(items)-20} more"

                row = _row(ws_reg, row, [
                    label if first_acct else "",
                    region if first_r else "",
                    rname if first_r else "",
                    svc, cnt, dtxt,
                ], alt=(row % 2 == 0))
                first_acct = False
                first_r = False

    _widths(ws_reg, {1:22, 2:18, 3:16, 4:28, 5:8, 6:80})
    ws_reg.auto_filter.ref = f"A2:{get_column_letter(rncols)}{row-1}"
    ws_reg.freeze_panes = "A3"

    # ==================================================================
    # SHEET 4: Service Quotas
    # ==================================================================
    ws_q = wb.create_sheet(title="Service Quotas")
    ws_q.sheet_properties.tabColor = "F39C12"
    qcols = ["Account", "Region", "Quota Name", "Value", "Adjustable", "Source"]
    qncols = len(qcols)

    row = _title(ws_q, 1, "SERVICE QUOTAS - EC2 (On-Demand Std, G&VT, Elastic IPs)", qncols)
    row = _headers(ws_q, row, qcols)

    for audit in audit_results:
        label = audit.get("account_label", "")
        quotas = audit.get("quotas") or {}
        first = True
        for rlabel, rquotas in quotas.items():
            for qname, qd in rquotas.items():
                if "error" in qd:
                    row = _row(ws_q, row, [
                        label if first else "", rlabel, qname, "ERROR", "", qd["error"][:60]
                    ])
                    ws_q.cell(row=row-1, column=4).font = ERR_FONT
                else:
                    row = _row(ws_q, row, [
                        label if first else "", rlabel, qname,
                        qd.get("value", "N/A"),
                        "Yes" if qd.get("adjustable") else "No",
                        qd.get("source", ""),
                    ], alt=(row % 2 == 0))
                first = False

    _widths(ws_q, {1:22, 2:28, 3:52, 4:12, 5:12, 6:18})
    ws_q.auto_filter.ref = f"A2:{get_column_letter(qncols)}{row-1}"
    ws_q.freeze_panes = "A3"

    # ==================================================================
    # SHEET 5: Organization Verification
    # ==================================================================
    ws_o = wb.create_sheet(title="Organization")
    ws_o.sheet_properties.tabColor = "8E44AD"
    ocols = ["Account Label", "Account ID", "Org ID", "Mgmt Account ID",
             "Mgmt Account Email", "Account Alias", "MFA Status", "Last Updated"]
    oncols = len(ocols)

    row = _title(ws_o, 1, "ORGANIZATION VERIFICATION", oncols)
    row = _headers(ws_o, row, ocols)

    for idx, audit in enumerate(audit_results):
        org = audit.get("organization") or {}
        row = _row(ws_o, row, [
            audit.get("account_label", ""),
            audit.get("account_id", ""),
            org.get("org_id", "N/A"),
            org.get("management_account_id", "N/A"),
            org.get("management_account_email", "N/A"),
            org.get("account_name", "N/A"),
            org.get("mfa_status", "N/A"),
            audit.get("last_updated", audit.get("audit_timestamp", "")),
        ], alt=(idx % 2 == 0))
        _status_font(ws_o, row-1, 7, org.get("mfa_status", ""))

    _widths(ws_o, {1:22, 2:16, 3:18, 4:18, 5:32, 6:20, 7:14, 8:22})
    ws_o.freeze_panes = "A3"

    # ==================================================================
    # SHEET 6: Checklist Status (6-step tracker)
    # ==================================================================
    ws_c = wb.create_sheet(title="Audit Checklist")
    ws_c.sheet_properties.tabColor = "27AE60"
    ccols = ["Account", "Step", "Checklist Item", "Status", "Details"]
    cncols = len(ccols)

    row = _title(ws_c, 1, "INTERNAL AUDIT CHECKLIST - 6 STEPS", cncols)
    row = _headers(ws_c, row, ccols)

    for audit in audit_results:
        label = audit.get("account_label", "")
        b = audit.get("billing") or {}
        p = audit.get("payment") or {}
        org = audit.get("organization") or {}
        rd = audit.get("regions") or {}
        q = audit.get("quotas") or {}

        ar = rd.get("regions", {}) if isinstance(rd, dict) else {}

        steps = [
            ("1", "Check Last Month Bill + Forecasted Cost",
             "Done" if b.get("last_month_total_cost", "N/A") != "N/A" else "Failed",
             f"Last Month: ${b.get('last_month_total_cost','N/A')} | "
             f"Forecast: ${b.get('forecasted_cost_current_month','N/A')}"),

            ("2", "Check Outstanding Balance / Bills Due",
             "Done" if p.get("outstanding_balance", "N/A") != "N/A" else "Failed",
             f"Outstanding: ${p.get('outstanding_balance','N/A')} | "
             f"Due: {p.get('payment_due','N/A')}"),

            ("3", "Check Regions & Services Running",
             "Done" if ar else "Failed",
             f"{len(ar)} region(s), "
             f"{rd.get('total_resources',0) if isinstance(rd,dict) else 0} resource(s)"),

            ("4", "Verify Organization (Org ID + Mgmt Email)",
             "Done" if org.get("org_id", "N/A") not in ("N/A", "") and not str(org.get("org_id","")).startswith("Error") else "Failed",
             f"Org: {org.get('org_id','N/A')} | Email: {org.get('management_account_email','N/A')}"),

            ("5", "Check Service Quotas (EC2 On-Demand, G&VT, EIPs)",
             "Done" if q else "Failed",
             f"Checked {len(q)} region(s)"),

            ("6", "Last Updated",
             "Done",
             audit.get("last_updated", audit.get("audit_timestamp", ""))),
        ]

        first = True
        for step_num, desc, status, detail in steps:
            row = _row(ws_c, row, [
                label if first else "", step_num, desc, status, detail,
            ], alt=(row % 2 == 0))
            _status_font(ws_c, row-1, 4, status)
            first = False

    _widths(ws_c, {1:22, 2:6, 3:48, 4:10, 5:60})
    ws_c.freeze_panes = "A3"

    # ── Save ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.getvalue()

    file_path = None
    if save_dir:
        dir_path = os.path.join(save_dir, str(ry), f"{rm:02d}")
        os.makedirs(dir_path, exist_ok=True)
        ts = now.strftime("%Y%m%d_%H%M%S")
        fname = f"AWS_Audit_{datetime(ry, rm, 1).strftime('%b_%Y')}_{ts}.xlsx"
        file_path = os.path.join(dir_path, fname)
        with open(file_path, "wb") as f:
            f.write(file_bytes)

    return file_bytes, file_path


# ══════════════════════════════════════════════════════════════════════
# OPTIMIZATION REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════════

CRIT_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid")
MED_FILL = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
LOW_FILL = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
CRIT_FONT = Font(name="Calibri", bold=True, size=10, color="C0392B")
HIGH_FONT_R = Font(name="Calibri", bold=True, size=10, color="E67E22")

SEVERITY_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
SEVERITY_FILL = {"Critical": CRIT_FILL, "High": HIGH_FILL, "Medium": MED_FILL, "Low": LOW_FILL}


def generate_optimization_report(optimization_results, save_dir=None):
    """Generate an Excel optimization report from scan results."""
    wb = Workbook()
    now = datetime.utcnow()

    # ==================================================================
    # SHEET 1: Executive Summary
    # ==================================================================
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = "E74C3C"

    ncols = 6
    row = _title(ws, 1, "AWS COST & SECURITY OPTIMIZATION REPORT", ncols)
    row = _title(ws, row,
                 f"Generated: {now.strftime('%Y-%m-%d %H:%M UTC')} | "
                 f"Accounts: {len(optimization_results)}", ncols)
    ws.cell(row=row-1, column=1).font = Font(name="Calibri", size=10, color=WHITE, italic=True)

    # Per-account summary
    row = _headers(ws, row, ["Account", "Total Findings", "Critical", "High",
                              "Est. Monthly Waste ($)", "Est. Annual Waste ($)"])

    grand_monthly = 0
    grand_findings = 0

    for opt in optimization_results:
        summary = opt.get("summary", {})
        by_sev = summary.get("by_severity", {})
        monthly = summary.get("estimated_monthly_waste", 0)
        grand_monthly += monthly
        grand_findings += summary.get("total_findings", 0)
        row = _row(ws, row, [
            opt.get("account_label", ""),
            summary.get("total_findings", 0),
            by_sev.get("Critical", 0),
            by_sev.get("High", 0),
            f"{monthly:,.2f}",
            f"{monthly * 12:,.2f}",
        ])
        if by_sev.get("Critical", 0) > 0:
            ws.cell(row=row-1, column=3).font = CRIT_FONT
        if by_sev.get("High", 0) > 0:
            ws.cell(row=row-1, column=4).font = HIGH_FONT_R

    # Totals row
    row += 1
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).border = BORDER
    ws.cell(row=row, column=1, value="TOTAL").font = BOLD
    ws.cell(row=row, column=2, value=grand_findings).font = BOLD
    ws.cell(row=row, column=5, value=f"{grand_monthly:,.2f}").font = BOLD
    ws.cell(row=row, column=6, value=f"{grand_monthly * 12:,.2f}").font = BOLD
    row += 2

    # Category breakdown
    c = ws.cell(row=row, column=1, value="Findings by Category")
    c.font = SUB_FONT; c.fill = SUB_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    all_cats = {}
    for opt in optimization_results:
        for cat, cnt in opt.get("summary", {}).get("by_category", {}).items():
            all_cats[cat] = all_cats.get(cat, 0) + cnt
    for cat, cnt in sorted(all_cats.items(), key=lambda x: -x[1]):
        row = _row(ws, row, [cat, cnt, "", "", "", ""])

    _widths(ws, {1: 25, 2: 16, 3: 10, 4: 10, 5: 22, 6: 22})
    ws.freeze_panes = "A4"

    # ==================================================================
    # SHEET 2: All Findings (detailed)
    # ==================================================================
    ws_f = wb.create_sheet(title="All Findings")
    ws_f.sheet_properties.tabColor = "F39C12"

    fcols = ["#", "Account", "Severity", "Category", "Resource Type", "Resource ID",
             "Region", "Issue", "Detail", "Est. Monthly Waste ($)", "Recommendation"]
    fncols = len(fcols)

    row = _title(ws_f, 1, "ALL OPTIMIZATION FINDINGS", fncols)
    row = _headers(ws_f, row, fcols)

    idx = 0
    for opt in optimization_results:
        label = opt.get("account_label", "")
        findings = opt.get("findings", [])
        sorted_findings = sorted(findings, key=lambda x: SEVERITY_ORDER.get(x.get("severity", "Low"), 99))

        for f in sorted_findings:
            idx += 1
            sev = f.get("severity", "Low")
            row = _row(ws_f, row, [
                idx, label, sev, f.get("category", ""),
                f.get("resource_type", ""), f.get("resource_id", ""),
                f.get("region", ""), f.get("issue", ""),
                f.get("detail", ""), f.get("estimated_monthly_waste", 0),
                f.get("recommendation", ""),
            ])
            # Color severity cell
            sev_fill = SEVERITY_FILL.get(sev)
            if sev_fill:
                ws_f.cell(row=row-1, column=3).fill = sev_fill
            if sev == "Critical":
                ws_f.cell(row=row-1, column=3).font = CRIT_FONT
            elif sev == "High":
                ws_f.cell(row=row-1, column=3).font = HIGH_FONT_R

    _widths(ws_f, {1: 5, 2: 20, 3: 10, 4: 14, 5: 16, 6: 28,
                   7: 16, 8: 35, 9: 55, 10: 18, 11: 55})
    ws_f.auto_filter.ref = f"A2:{get_column_letter(fncols)}{max(row-1, 3)}"
    ws_f.freeze_panes = "A3"

    # ==================================================================
    # SHEET 3: Cost Savings Summary
    # ==================================================================
    ws_cost = wb.create_sheet(title="Cost Savings")
    ws_cost.sheet_properties.tabColor = "27AE60"

    ccols = ["Account", "Issue Type", "Count", "Est. Monthly Savings ($)", "Est. Annual Savings ($)"]
    cncols = len(ccols)

    row = _title(ws_cost, 1, "POTENTIAL COST SAVINGS BREAKDOWN", cncols)
    row = _headers(ws_cost, row, ccols)

    for opt in optimization_results:
        label = opt.get("account_label", "")
        findings = opt.get("findings", [])
        # Group by issue
        issue_groups = {}
        for f in findings:
            if f.get("estimated_monthly_waste", 0) > 0:
                issue = f.get("issue", "Other")
                if issue not in issue_groups:
                    issue_groups[issue] = {"count": 0, "monthly": 0}
                issue_groups[issue]["count"] += 1
                issue_groups[issue]["monthly"] += f.get("estimated_monthly_waste", 0)

        first = True
        for issue, data in sorted(issue_groups.items(), key=lambda x: -x[1]["monthly"]):
            row = _row(ws_cost, row, [
                label if first else "",
                issue, data["count"],
                f"{data['monthly']:,.2f}",
                f"{data['monthly'] * 12:,.2f}",
            ], alt=(row % 2 == 0))
            first = False

        if not issue_groups:
            row = _row(ws_cost, row, [label, "No cost savings found", 0, "0.00", "0.00"])

    _widths(ws_cost, {1: 22, 2: 42, 3: 8, 4: 22, 5: 22})
    ws_cost.freeze_panes = "A3"

    # ==================================================================
    # SHEET 4: Security Findings
    # ==================================================================
    ws_sec = wb.create_sheet(title="Security Findings")
    ws_sec.sheet_properties.tabColor = "8E44AD"

    scols = ["Account", "Severity", "Resource Type", "Resource ID",
             "Region", "Issue", "Recommendation"]
    sncols = len(scols)

    row = _title(ws_sec, 1, "SECURITY FINDINGS", sncols)
    row = _headers(ws_sec, row, scols)

    has_sec = False
    for opt in optimization_results:
        label = opt.get("account_label", "")
        findings = [f for f in opt.get("findings", []) if f.get("category") == "Security"]
        sorted_findings = sorted(findings, key=lambda x: SEVERITY_ORDER.get(x.get("severity", "Low"), 99))

        for f in sorted_findings:
            has_sec = True
            sev = f.get("severity", "Low")
            row = _row(ws_sec, row, [
                label, sev, f.get("resource_type", ""),
                f.get("resource_id", ""), f.get("region", ""),
                f.get("issue", ""), f.get("recommendation", ""),
            ])
            sev_fill = SEVERITY_FILL.get(sev)
            if sev_fill:
                ws_sec.cell(row=row-1, column=2).fill = sev_fill

    if not has_sec:
        row = _row(ws_sec, row, ["", "", "", "", "", "No security issues found", ""])
        ws_sec.cell(row=row-1, column=6).font = OK_FONT

    _widths(ws_sec, {1: 22, 2: 10, 3: 16, 4: 28, 5: 16, 6: 42, 7: 55})
    ws_sec.freeze_panes = "A3"

    # ── Save ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.getvalue()

    file_path = None
    if save_dir:
        dir_path = os.path.join(save_dir, "optimization")
        os.makedirs(dir_path, exist_ok=True)
        ts = now.strftime("%Y%m%d_%H%M%S")
        fname = f"AWS_Optimization_{ts}.xlsx"
        file_path = os.path.join(dir_path, fname)
        with open(file_path, "wb") as f:
            f.write(file_bytes)

    return file_bytes, file_path
